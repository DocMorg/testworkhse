from openpyxl import load_workbook
from sys import argv
import psycopg2
import argparse
import re
import datetime


def create_table(sheet, table_name, flag=False):
    # Проверим, не существует ли уже данная таблица в бд:
    # Подключение к бд и получение курсора
    conn = psycopg2.connect(dbname='postgres', user='root',
                            password='toor', host='localhost')
    cursor = conn.cursor()
    conn.autocommit = True
    query = "select tablename from pg_catalog.pg_tables where schemaname != 'information_schema' \
             and schemaname != 'pg_catalog' and tablename = '" + str(table_name) + "';"
    cursor.execute(query)
    # Если таблица существует - выходим, иначе - просто идём дальше
    if len(cursor.fetchall()) > 0:
        exit('Table already exists in database, use argument -upd instead of -c')
    # Считаем названия столбцов в массив и определим их количество
    col_names = []
    i = 1
    while sheet.cell(1, i).value is not None:
        col_names.append(sheet.cell(1, i).value)
        i += 1
    max_col = i-1
    # Проверка на пустую таблицу
    if max_col < 1:
        exit('Table given is empty. Enter any data and try again.')
    # Найдем столбец с самой большой длиной, из-за возможных пропущенных значений
    j = 1
    max_line = -1
    for n in [1, max_col]:
        while sheet.cell(j, n).value:
            j += 1
        if j > max_line:
            max_line = j-2  # differ 2 because of last j+=1 and also we don't need names line
    # Проверка на пустую таблицу
    if max_line < 1:
        exit('Table given is empty. Enter any data and try again.')
    # Дальше - обработка данных с форматированием
    data = [[''] * max_line for n in range(max_col)]
    types = []
    for n in range(1, max_col+1):
        flag1 = False
        flag2 = False
        flag3 = False
        flag4 = False
        for j in range(2, max_line+2):
            if len(str(sheet.cell(j, n).value)) == 0 or sheet.cell(j, n).value is None:
                k = 'NULL'
            else:
                k = sheet.cell(j, n).value
                if isinstance(k, datetime.datetime):
                    flag1 = True
                else:
                    try:
                        if not k % 1:
                            raise ValueError
                        else:
                            flag2 = True
                    except (ValueError, TypeError):
                        try:
                            k = int(k)
                            flag3 = True
                        except ValueError:
                            try:
                                k = str(k)
                                flag4 = True
                            except ValueError:
                                pass
            data[n-1][j-2] = k
        # Определим формат столбца и запишем его в список
        if flag1 and not (flag2 or flag3 or flag4):
            types.append('timestamp')
        else:
            if flag2 and not (flag1 or flag3 or flag4):
                types.append('numeric')
            else:
                if flag3 and not (flag1 or flag2 or flag4):
                        types.append('bigint')
                else:
                    types.append('text')
    query = ''
    for i, j in zip(col_names, types):
        query += str(i) + ' ' + str(j) + ', '
    sqlcreatetable = 'create table ' + str(table_name) + ' (' + query[:-2] + ');'
    cursor.execute(sqlcreatetable)
    add = ''
    for i in col_names:
        add += str(i) + ', '
    values = ''
    for j in range(0, max_line):
        string = '('
        for i in range(0, max_col):
            if types[i] == 'timestamp' and data[i][j] == 'NULL':
                string += 'NULL, '
            else:
                if types[i] == 'text' or types[i] == 'timestamp':
                    string += "'" + str(data[i][j]) + "', "
                else:
                    string += str(data[i][j]) + ', '
        values += string[:-2] + '), '
    sqladd = 'insert into ' + str(table_name) + ' (' + add[:-2] + ') values' + values[:-2] + ';'
    cursor.execute(sqladd)
    cursor.close()
    conn.close()
    if flag:
        return col_names, types
    else:
        exit('Table created successfully and filled with data')


def update_table(sheet, table_name):
    # Проверим, не существует ли уже данная таблица в бд:
    # Подключение к бд и получение курсора
    conn = psycopg2.connect(dbname='postgres', user='root',
                            password='toor', host='localhost')
    conn.autocommit = True
    cursor = conn.cursor()
    check = "select tablename from pg_catalog.pg_tables where schemaname != 'information_schema' \
                 and schemaname != 'pg_catalog' and tablename like '" + str(table_name) + "' ESCAPE '/';"
    cursor.execute(check)
    # Если таблица существует - выходим, иначе - просто идём дальше
    if len(cursor.fetchall()) == 0:
        exit('Table does not exists in database, use argument -c instead of -upd')
    new_table_name = 'new____' + str(table_name)
    col_names, types = create_table(sheet, new_table_name, True)
    check1 = "SELECT data_type FROM information_schema.columns where table_name = '" + str(table_name) + "';"
    cursor.execute(check1)
    new_types = cursor.fetchall()
    new_types = [(type1[0]) for type1 in new_types]
    for i in range(len(types)):
        if types[i] != new_types[i]:
            query1 = 'ALTER TABLE ' + str(table_name) + ' ALTER COLUMN ' + \
                     str(col_names[i]) + ' TYPE text;'
            cursor.execute(query1)
    query = ''
    for i in col_names:
        query += str(i) + ', '
    query = query[:-2]
    query = 'insert into ' + str(table_name) + ' (' + query + ') '\
            'select ' + query + ' from ' + new_table_name + ' where '\
            'not exists(select 1 from ' + str(table_name) + ' where '\
            + col_names[0] + ' = ' + new_table_name + '.' + col_names[0] + \
            ' and ' + col_names[1] + ' = ' + new_table_name + '.' + col_names[1] +\
            ' ); ' + 'drop table if exists ' + new_table_name + ';'
    cursor.execute(query)
    cursor.close()
    conn.close()
    exit('Data successfully updated')


def add_index(index, table_name):
    conn = psycopg2.connect(dbname='postgres', user='root',
                            password='toor', host='localhost')
    conn.autocommit = True
    cursor = conn.cursor()
    query = "select column_name from INFORMATION_SCHEMA.columns where table_name = '" + str(table_name) + "';"
    cursor.execute(query)
    rows = cursor.fetchall()
    row_name = ''
    try:
        row_name = rows[index-1][0]
    except IndexError:
        exit('Column with given number does not exist. Index out of range error.')
    testquery = "select t.relname as table_name, i.relname as index_name, a.attname as column_name from\
                pg_class t, pg_class i, pg_index ix, pg_attribute a where t.oid = ix.indrelid and i.oid \
                = ix.indexrelid and a.attrelid = t.oid and a.attnum = ANY(ix.indkey) and t.relkind = 'r' \
                and t.relname like '" + str(table_name) + "' order by t.relname, i.relname;"
    cursor.execute(testquery)
    result = cursor.fetchall()[0]
    if result[1] == "id" + str(index) or result[2] == row_name:
        exit('Column index with the name already exists or column has an index already')
    # query = "create unique index id" + str(index) + " ON " + str(table_name) + "(" + str(row_name) + ");"
    # cursor.execute(query)
    cursor.close()
    conn.close()
    exit('Index added successfully to the selected column')


def main(args):
    # Создадим парсер и проверим входные аргументы
    excel = ''
    if re.fullmatch(r'((((?<!\w)[A-Z,a-z]:)|(\.{1,2}\\))([^\b%\/\|:\n\"]'
                    r'*))|(\2([^%\/\|:\n\"]*))|((?<!\w)(\.{1,2})?(?<!\/)'
                    r'(\/((\\\b)|[^ \b%\|:\n\"\\\/])+)+\/?)',
                    args.excel[0]):
        excel = args.excel[0]
    else:
        exit('Enter valid path to file')
    if args.update and args.create:
        exit('Please, select only ONE of the flag arguments: create or update')
    # Также проверим, нужного ли формата файл и можем ли его открыть,
    # если нет - выдаем ошибку и выходим из программы.
    try:
        wb = load_workbook(excel, data_only=True)
    except FileNotFoundError:
        print('Error FileNot Found: Enter file path carefully once more')
        return 1
    # Выберем активный лист. По дефолту, программа работает только с ним.
    # Для работы с другими листами нужно поменять следующую строку для выбора нужного.
    table_name = wb.sheetnames[wb._active_sheet_index]
    sheet = wb.active
    if args.create:
        create_table(sheet, table_name)
    if args.update:
        update_table(sheet, table_name)
    if args.add_index:
        index_col = args.add_index[0]
        add_index(index_col, table_name)
    if not(args.create or args.update or args.add_index):
        exit('No action specified')


if __name__ == '__main__':
    parser = argparse.ArgumentParser()
    parser.add_argument('excel', type=str, action='store',
                        help='path to xlsx file', nargs=1)
    parser.add_argument('-c', '--create', action='store_true',
                        help='flag if entered file is to add new table to database')
    parser.add_argument('-ai', '--add_index', action='store', type=int,
                        help="adds index to the entered column by its number, if table and column exists", nargs=1)
    parser.add_argument('-upd', '--update', action='store_true',
                        help="flag if entered file is to update the existing table")
    if len(argv) < 1:
        exit('No arguments entered')
    elif len(argv) < 2:
        exit('ERROR! You must specify path to file')
    args = parser.parse_args()
    main(args)

