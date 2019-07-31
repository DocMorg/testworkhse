from openpyxl import load_workbook
from sys import argv
import psycopg2
import argparse
import re
import datetime


def create_parser():
    parser = argparse.ArgumentParser()
    parser.add_argument('excel', type=str, action='store',
                        help='path to xlsx file', nargs=1)
    parser.add_argument('-ai', '--add_index', action='store', type=int,
                        help="adds index to the entered column by its number, if table and column exists", nargs=1)
    parser.add_argument('-upd', '--update', action='store',
                        help="flag if entered file is to update the existing table", nargs=1)
    print(argv)
    if len(argv) == 1:
        print('ERROR! You must specify path to file')
        return None
    else:
        return parser


def main():
    # Создадим парсер и проверим входные аргументы
    parser = create_parser()
    if parser is None:
        return 1
    args = parser.parse_args()
    if re.fullmatch(r'((((?<!\w)[A-Z,a-z]:)|(\.{1,2}\\))([^\b%\/\|:\n\"]'
                    r'*))|(\2([^%\/\|:\n\"]*))|((?<!\w)(\.{1,2})?(?<!\/)'
                    r'(\/((\\\b)|[^ \b%\|:\n\"\\\/])+)+\/?)',
                    args.excel[0]):
        excel = args.excel[0]
    else:
        return 1
    if args.add_index:
        index_col = args.add_index[0]
    # Также проверим, нужного ли формата файл и можем ли его открыть,
    # если нет - выдаем ошибку и выходим из программы.
    try:
        wb = load_workbook(excel, data_only=True)
    except FileNotFoundError:
        print('Error FileNot Found: Enter file path carefully once more')
        return 1
    # Выберем активный лист. По дефолту, программа работает только с ним.
    # Для работы с другими листами нужно поменять следующую строку для выбора нужного.
    table_name = wb.sheetnames[wb._active_sheet_index]
    sheet = wb.active
    # Считаем названия столбцов в массив и определим их количество
    col_names = []
    i = 1
    while sheet.cell(1, i).value is not None:
        col_names.append(sheet.cell(1, i).value)
        i += 1
    max_col = i-1
    # Найдем столбец с самой большой длиной, из-за возможных пропущенных значений
    j = 1
    max_line = -1
    for n in [1, max_col]:
        while sheet.cell(j, n).value:
            j += 1
        if j > max_line:
            max_line = j-1
    # Дальше - обработка данных с форматированием
    data = [[''] * max_line for n in range(max_col)]
    types = []
    for n in range(1, max_col+1):
        flag1 = True
        flag2 = True
        flag3 = True
        flag4 = True
        for j in range(2, max_line+2):
            if sheet.cell(j, n).value is None:
                k = 'NULL'
            else:
                k = sheet.cell(j, n).value
                if type(k) is datetime.datetime:
                    pass
                else:
                    flag1 = False
                    try:
                        if not k % 1:
                            raise ValueError
                    except (ValueError, TypeError):
                        flag2 = False
                        flag3 = False
                        try:
                            k = int(k)
                        except ValueError:
                            flag4 = False
                            try:
                                k = str(k)
                            except ValueError:
                                pass
            data[n-1][j-2] = k
        # Определим формат столбца и запишем его в список
        if flag1 and not flag2 & flag3 & flag4:
            types.append('timestamp')
        else:
            if flag2 or flag4 and not flag1 & flag3:
                types.append('numeric')
            else:
                if flag3 and not flag1 & flag2 & flag4:
                        types.append('bigint')
                else:
                    types.append('text')

    conn = psycopg2.connect(dbname='postgres', user='root',
                            password='toor', host='localhost')
    cursor = conn.cursor()
    query = ''
    for i, j in zip(col_names, types):
        query += str(i) + ' ' + str(j) + ', '
    # sqlcreatetable = 'create table ' + str(table_name) + ' (' + query[:-2] + ');'
    # cursor.execute(sqlcreatetable)
    conn.commit()
    add = ''
    for i in col_names:
        add += str(i) + ', '
    values = ''
    for j in range(0, max_line):
        string = '('
        for i in range(0, max_col):
            if types[i] == 'text' or types[i] == 'timestamp':
                string += "'" + str(data[i][j]) + "'" + ', '
            else:
                string += str(data[i][j]) + ', '
        values += string[:-2] + '), '
    # sqladd = 'insert into ' + str(table_name) + ' (' + add[:-2] + ') values' + values[:-2] + ';'
    # cursor.execute(sqladd)
    conn.commit()
    cursor.close()
    conn.close()


if __name__ == '__main__':
    main()
